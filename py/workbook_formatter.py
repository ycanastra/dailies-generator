import json

from openpyxl.styles import Alignment, Font, Border
from openpyxl.styles.numbers import FORMAT_DATE_TIME1
from openpyxl.styles.borders import Side
from openpyxl.styles.fills import PatternFill

class WorkbookFormatter:
	def __init__(self, ws, group):
		self.ws = ws
		self.group = group
		noSide = Side(border_style=None)
		thinSide = Side(border_style='thin', color='000000')
		self.noBorder = Border(top=noSide, bottom=noSide, left=noSide, right=noSide)
		self.outsideBorder = Border(top=Side(border_style='thin', color='000000'),
									  bottom=Side(border_style='thin', color='000000'),
									  left=Side(border_style='thin', color='000000'),
									  right=Side(border_style='thin', color='000000'))

		self.sideBorder = Border(top=Side(border_style=None, color='000000'),
								   bottom=Side(border_style=None, color='000000'),
								   left=Side(border_style='thin', color='000000'),
								   right=Side(border_style='thin', color='000000'))

		self.bottomBorder = Border(top=Side(border_style=None, color='000000'),
									 bottom=Side(border_style='thin', color='000000'),
									 left=Side(border_style='thin', color='000000'),
									 right=Side(border_style='thin', color='000000'))

		self.topBorder = Border(top=Side(border_style='thin', color='000000'),
								  bottom=Side(border_style=None, color='000000'),
								  left=Side(border_style='thin', color='000000'),
								  right=Side(border_style='thin', color='000000'))


		self.whiteFill = PatternFill(fill_type='solid', start_color='FFFFFF')

		self.yellowFill = PatternFill(fill_type='solid', start_color='FFFF99')

		self.grayFill = PatternFill(fill_type='solid', start_color='404040')

		self.lightgrayFill = PatternFill(fill_type='solid', start_color='eeeeee')

		self.centerAlignment = Alignment(horizontal='center', vertical='center',
										   wrap_text=True, shrink_to_fit=False)

		self.topAlignment = Alignment(horizontal='center', vertical='top',
										wrap_text=True, shrink_to_fit=False)

		self.getFormatData()

	def getFormatData(self):
		group = self.group
		path = './data/group_data.json'

		file = open(path)
		data = json.load(file)
		data = data[group]['Worksheet Format']

		self.titleCellHeight = data['Title']['cell_height']
		self.titleFont = data['Title']['font']
		self.titleFontSize = data['Title']['font_size']
		self.titleFontBold = data['Title']['font_bold']

		self.headerCellHeight = data['Header']['cell_height']
		self.headerFont = data['Header']['font']
		self.headerFontSize = data['Header']['font_size']
		self.headerFontBold = data['Header']['font_bold']

		self.timeCellWidth = data['Time Column']['cell_width']
		self.timeFont = data['Time Column']['font']
		self.timeFontSize = data['Time Column']['font_size']
		self.timeFontBold = data['Time Column']['font_bold']

		self.labCellWidth = data['Lab Column']['cell_width']
		self.labFont = data['Lab Column']['font']
		self.labFontSize = data['Lab Column']['font_size']
		self.labFontBold = data['Lab Column']['font_bold']

		self.employeeCellWidth = data['Employee Column']['cell_width']
		self.employeeFont = data['Employee Column']['font']
		self.employeeFontSize = data['Employee Column']['font_size']
		self.employeeFontBold = data['Employee Column']['font_bold']

		self.notesCellWidth = data['Notes Column']['cell_width']

		self.cellHeight = data['cell_height']

	# TODO fix so it works when sCol!=eCol
	def setBorder(self, sRow, eRow, sCol, eCol):
		ws = self.ws
		if sRow == eRow:
			ws.cell(row=sRow, column=sCol).border = self.outsideBorder
		else:
			ws.cell(row=sRow, column=sCol).border = self.topBorder
			ws.cell(row=eRow, column=sCol).border = self.bottomBorder

			for i in range(sRow + 1, eRow):
				ws.cell(row=i, column=sCol).border = self.sideBorder

	def setSolidFill(self, sRow, eRow, sCol, eCol, color):
		ws = self.ws
		patternFill = PatternFill(fill_type='solid', start_color=color)
		for i in range(sRow, eRow + 1):
			for j in range(sCol, eCol + 1):
				ws.cell(row=i, column=j).fill = patternFill

	def setFontSize(self, sRow, eRow, sCol, eCol, fontSize):
		ws = self.ws
		for i in range(sRow, eRow + 1):
			for j in range(sCol, eCol + 1):
				ws.cell(row=i, column=j).font = Font(size=fontSize)

	def setCenterAlignment(self, sRow, eRow, sCol, eCol):
		ws = self.ws
		for i in range(sRow, eRow + 1):
			for j in range(sCol, eCol + 1):
				ws.cell(row=i, column=j).alignment = self.centerAlignment

	def setTopAlignment(self, sRow, eRow, sCol, eCol):
		ws = self.ws
		for i in range(sRow, eRow + 1):
			for j in range(sCol, eCol + 1):
				ws.cell(row=i, column=j).alignment = self.topAlignment

	def setMerge(self, sRow, eRow, sCol, eCol):
		ws = self.ws
		ws.merge_cells(start_row=sRow, start_column=sCol, end_row=eRow, end_column=eCol)
