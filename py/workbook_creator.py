import openpyxl
import datetime
import json
import os

from workbook_formatter import WorkbookFormatter

from collections import OrderedDict

from openpyxl.styles import Alignment, Font, Border
from openpyxl.styles.numbers import FORMAT_DATE_TIME1
from openpyxl.styles.borders import Side
from openpyxl.styles.fills import PatternFill
from openpyxl.utils import get_column_letter

class WorkbookCreator:
	def __init__(self, soup, name, group, date, scheduleEvents, workerEvents):
		self.__soup = soup
		self.__name = name
		self.__group = group
		self.__date = date
		self.__scheduleEvents = scheduleEvents
		self.__workerEvents = workerEvents

		self.__wb = openpyxl.Workbook()
		self.__ws = self.__wb.worksheets[0]
		self.__ws.title = group

		self.__wf = WorkbookFormatter(self.__ws, self.__group)

	def generateWorkbook(self):
		group = self.__group
		labLocations = self.getLabLocations()
		employeeLocations = self.getEmployeeLocations()

		currentColumn = 1
		currentRow = 2

		self.createTimeAxis(currentRow, currentColumn)

		currentColumn += 1

		for key, value in labLocations.iteritems():
			self.createLabColumn(currentRow, currentColumn, (key, value))
			currentColumn += 1

		for key, value in employeeLocations.iteritems():
			self.createEmployeeColumn(currentRow, currentColumn, (key, value))
			currentColumn += 1

		self.createTimeAxis(currentRow, currentColumn)

		currentColumn += 1

		self.createNotesColumn(currentRow, currentColumn)

		currentRow = 1
		currentColumn = 1

		self.createHeader(currentRow, currentColumn)

		self.__wb.save(group + '.xlsx')

	def createLabColumn(self, row, col, labNamePair):
		ws = self.__ws
		wf = self.__wf
		fontSize = wf.fontSize
		cellWidth = wf.cellWidth

		labNameKey = labNamePair[0] # Lab name that shows up on online labschedule
		labNameValue = labNamePair[1] # Lab name to display on dailies

		scheduleEvents = self.__scheduleEvents

		ws.column_dimensions[get_column_letter(col)].width = cellWidth

		# LabName Title
		currentCell = ws.cell(row=row, column=col)
		currentCell.value = labNameValue
		currentCell.alignment = wf.centerAlignment
		currentCell.font = Font(size=fontSize, bold=True)
		currentCell.border = wf.outsideBorder

		scheduleEvents = [x for x in scheduleEvents if x.getLocation() == labNameKey]

		for item in scheduleEvents:
			if item.getEventName() == labNameKey: # Skip useless scheduleEvent
				continue

			eventName = item.getEventName()
			eventTimeString = item.getTimeString()
			eventTeacher = shortenTeacherName(item.getEventTeacher())

			startTime = item.getStartTime()
			endTime = item.getEndTime()

			startTimeInt = int(startTime.hour) + int(startTime.minute)/60.0
			endTimeInt = int(endTime.hour) + int(endTime.minute)/60.0

			startRow = int((startTimeInt - 8)*2 + row + 1)
			endRow = int((endTimeInt - 8)*2 + row)

			wf.setFontSize(startRow, endRow, col, col, fontSize)
			wf.setCenterAlignment(startRow, endRow, col, col)

			if item.getEventName() == 'OPEN':
				# self.setSolidFill(startRow, endRow, col, 'FFFFFF') # White

				# patternFill = PatternFill(fill_type='solid', start_color=color)
				for i in range(startRow, endRow + 1):
					if (i + 1)%4 < 2:
						ws.cell(row=i, column=col).fill = wf.whiteFill
					else:
						ws.cell(row=i, column=col).fill = wf.lightgrayFill

				wf.setBorder(startRow, endRow, col, col)
				continue
			elif item.getEventName() == 'CLOSED':
				wf.setSolidFill(startRow, endRow, col, col, '404040') # Gray
				wf.setBorder(startRow, endRow, col, col)
				continue
			else: # This is an actual class
				wf.setSolidFill(startRow, endRow, col, col, 'FFFFFF') # Yellow
				wf.setBorder(startRow, endRow, col, col)

				# Making sure the event takes up enough cells to merge properly
				if endRow - startRow >= 2:
					wf.setMerge(startRow, endRow - 2, col, col)

				ws.cell(row=startRow, column=col).value = eventName

				if startRow != endRow: # If the event takes up more than one cell
					ws.cell(row=endRow, column=col).value = eventTimeString

					for i in range(startRow + 1, endRow):
						if i == endRow - 1:
							ws.cell(row=i, column=col).value = eventTeacher

	def createEmployeeColumn(self, row, col, employeeLocationPair):
		ws = self.__ws
		wf = self.__wf
		fontSize = wf.fontSize
		cellWidth = wf.cellWidth
		workerEvents = self.__workerEvents
		day = (self.__date.weekday() + 1)%7

		# Employee location names that are contained in ./data/employee_shifts/*
		employeeLocationKey = employeeLocationPair[0]
		# Employee location name that is displayed on the dailies
		employeeLocationValue = employeeLocationPair[1]

		ws.column_dimensions[get_column_letter(col)].width = cellWidth

		# employeeLocation Title
		currentCell = ws.cell(row=row, column=col)
		currentCell.value = employeeLocationValue
		currentCell.alignment = wf.centerAlignment
		currentCell.font = Font(size=fontSize, bold=True)
		currentCell.border = wf.outsideBorder

		workerEvents = [x for x in workerEvents if x.getDay() == day]
		workerEvents = [x for x in workerEvents if x.getLocation() == employeeLocationKey]

		for item in workerEvents:
			name = item.getName()
			timeString = item.getTimeString()

			startTime = item.getStartTime()
			endTime = item.getEndTime()

			startTimeInt = int(startTime.hour) + int(startTime.minute)/60.0
			endTimeInt = int(endTime.hour) + int(endTime.minute)/60.0

			startRow = int((startTimeInt - 8)*2 + row + 1)
			endRow = int((endTimeInt - 8)*2 + row)

			ws.cell(row=startRow, column=col).value = name
			ws.cell(row=endRow, column=col).value = timeString

			wf.setBorder(startRow, endRow, col, col)
			wf.setCenterAlignment(startRow, endRow, col, col)
			wf.setFontSize(startRow, endRow, col, col, fontSize)
			wf.setSolidFill(startRow, endRow, col, col, 'FFFFFF') # White

			wf.setBorder(endRow + 1, ws.max_row, col, col)
			wf.setCenterAlignment(endRow + 1, ws.max_row, col, col)
			wf.setFontSize(endRow + 1, ws.max_row, col, col, fontSize)
			wf.setSolidFill(endRow + 1, ws.max_row, col, col, '404040') # Gray

	def createTimeAxis(self, row, col):
		ws = self.__ws
		wf = self.__wf
		timeCellWidth = wf.timeCellWidth
		cellHeight = wf.cellHeight
		fontSize = wf.fontSize

		self.addTimeTitle(row, col)

		ws.column_dimensions[get_column_letter(col)].width = timeCellWidth
		wf.setTopAlignment(row + 1, row + 29, col, col)

		for i in xrange(0, 29):
			currentRow = row + i + 1
			currentCell = ws.cell(row=currentRow, column=col)

			ws.row_dimensions[currentRow].height = cellHeight

			if i%2 == 0:
				currentCell.value = datetime.datetime.strptime(str(i/2 + 8), '%H')

			currentCell.number_format = FORMAT_DATE_TIME1
			currentCell.font = Font(size=fontSize)
			if (currentRow + 1)%4 < 2:
				currentCell.fill = wf.whiteFill
			else:
				currentCell.fill = wf.lightgrayFill

			if i == 0:
				currentCell.border = wf.topBorder
			elif i == 28:
				currentCell.border = wf.bottomBorder
			else:
				currentCell.border = wf.sideBorder

	def createNotesColumn(self, row, col):
		ws = self.__ws
		wf = self.__wf
		cellHeight = wf.cellHeight
		fontSize = wf.fontSize

		self.addNotesTitle(row, col)

		# ws.column_dimensions[get_column_letter(col)].width = timeCellWidth

		for i in xrange(0, 29):
			currentRow = row + i + 1
			currentCell = ws.cell(row=currentRow, column=col)

			ws.row_dimensions[currentRow].height = cellHeight

			currentCell.font = Font(size=fontSize)

			if (currentRow + 1)%4 < 2:
				currentCell.fill = wf.whiteFill
			else:
				currentCell.fill = wf.lightgrayFill

			if i == 0:
				currentCell.border = wf.topBorder
			elif i == 28:
				currentCell.border = wf.bottomBorder
			else:
				currentCell.border = wf.sideBorder

	def createHeader(self, row, col):
		ws = self.__ws
		wf = self.__wf
		dateString = self.__date.strftime('%m/%d/%Y')
		name = 'Name: ' + self.__name
		group = self.__group

		headerHeight = wf.headerHeight
		headerFontSize = wf.headerFontSize

		title = 'CSSC ' + group

		maxCol = ws.max_column
		maxTitleCol = maxCol*1/3
		maxDateCol = maxCol*2/3

		wf.setMerge(row, row, col, maxTitleCol)
		wf.setMerge(row, row, maxTitleCol + 1, maxDateCol)
		wf.setMerge(row, row, maxDateCol + 1, maxCol)

		ws.cell(row=row, column=col).value = title
		ws.cell(row=row, column=maxTitleCol + 1).value = dateString
		ws.cell(row=row, column=maxDateCol + 1).value = name

		ws.row_dimensions[row].height = headerHeight

		ws.cell(row=row, column=col).font = Font(size=headerFontSize, bold=True)
		ws.cell(row=row, column=maxTitleCol + 1).font = Font(size=headerFontSize, bold=True)
		ws.cell(row=row, column=maxDateCol + 1).font = Font(size=headerFontSize, bold=True)

		wf.setCenterAlignment(row, row, col, maxDateCol + 1)


	def getEmployeeLocations(self):
		group = self.__group
		path = './data/group_data.json'

		file = open(path)
		data = json.load(file, object_pairs_hook=OrderedDict)

		return data[group]['Employee Locations']

	def getLabLocations(self):
		group = self.__group
		path = './data/group_data.json'

		file = open(path)
		data = json.load(file, object_pairs_hook=OrderedDict)

		return data[group]['Labs']


	def getFormatData(self):
		group = self.__group
		path = './data/group_data.json'

		file = open(path)
		data = json.load(file)
		data = data[group]['Worksheet Format']

		self.__cellHeight = data['cell_height']
		self.__cellWidth = data['cell_width']
		self.__fontSize = data['font_size']
		self.__timeCellWidth = data['time_cell_width']
		self.__headerHeight = data['header_height']
		self.__headerFontSize = data['header_font_size']

	def addTimeTitle(self, row, col):
		ws = self.__ws
		wf = self.__wf
		fontSize = wf.fontSize
		cellHeight = wf.cellHeight
		group = self.__group

		ws.row_dimensions[row].height = cellHeight

		timeCell = ws.cell(row=row, column=col)
		timeCell.value = 'Time'
		timeCell.alignment = wf.centerAlignment
		timeCell.font = Font(size=fontSize, bold=True)
		timeCell.border = wf.outsideBorder


	def addNotesTitle(self, row, col):
		ws = self.__ws
		wf = self.__wf
		fontSize = wf.fontSize

		notesCell = ws.cell(row=row, column=col)
		notesCell.value = 'Notes'
		notesCell.alignment = wf.centerAlignment
		notesCell.font = Font(size=fontSize, bold=True)
		notesCell.border = wf.outsideBorder

def shortenTeacherName(teacherName):
	if teacherName == 'To Be Determined':
		return 'TBD'
	elif teacherName == None:
		return None
	else:
		return teacherName
